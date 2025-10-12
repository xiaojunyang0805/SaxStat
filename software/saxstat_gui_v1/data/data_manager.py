"""
Data Manager

Handles data collection, storage, and export for experiments.
Uses pandas for efficient data handling.
"""

import pandas as pd
from typing import Dict, List, Optional, Any
from pathlib import Path
from datetime import datetime


class DataManager:
    """
    Manages experiment data collection and export.

    Features:
    - Pandas DataFrame storage
    - Multiple export formats (CSV, Excel, JSON)
    - Metadata preservation
    - Incremental data addition
    """

    def __init__(self):
        self.data: Optional[pd.DataFrame] = None
        self.metadata: Dict[str, Any] = {}
        self._reset()

    def _reset(self):
        """Reset data storage."""
        self.data = pd.DataFrame()
        self.metadata = {
            'experiment_name': '',
            'start_time': None,
            'end_time': None,
            'parameters': {},
            'notes': ''
        }

    # Data collection

    def start_collection(self, experiment_name: str, parameters: Dict[str, Any]):
        """
        Start a new data collection session.

        Args:
            experiment_name: Name of the experiment
            parameters: Experiment parameters
        """
        self._reset()
        self.metadata['experiment_name'] = experiment_name
        self.metadata['start_time'] = datetime.now()
        self.metadata['parameters'] = parameters.copy()

    def add_data_point(self, data_point: Dict[str, float]):
        """
        Add a single data point.

        Args:
            data_point: Dictionary with data values
                e.g., {'time': 1.5, 'voltage': 0.5, 'current': 10.2}
        """
        new_row = pd.DataFrame([data_point])
        self.data = pd.concat([self.data, new_row], ignore_index=True)

    def add_data_batch(self, data_points: List[Dict[str, float]]):
        """
        Add multiple data points at once.

        Args:
            data_points: List of data point dictionaries
        """
        new_data = pd.DataFrame(data_points)
        self.data = pd.concat([self.data, new_data], ignore_index=True)

    def complete_collection(self):
        """Mark data collection as complete."""
        self.metadata['end_time'] = datetime.now()

    # Data access

    def get_data(self) -> pd.DataFrame:
        """
        Get the collected data as a pandas DataFrame.

        Returns:
            pd.DataFrame: Data frame with all collected data
        """
        return self.data.copy()

    def get_column(self, column_name: str) -> List[float]:
        """
        Get a specific data column.

        Args:
            column_name: Name of the column

        Returns:
            list: Column values
        """
        if column_name in self.data.columns:
            return self.data[column_name].tolist()
        return []

    def get_metadata(self) -> Dict[str, Any]:
        """Get experiment metadata."""
        return self.metadata.copy()

    def is_empty(self) -> bool:
        """Check if data is empty."""
        return self.data.empty

    # Data export

    def export_csv(self, filepath: Path, include_metadata: bool = True):
        """
        Export data to CSV format.

        Args:
            filepath: Output file path
            include_metadata: If True, prepend metadata as comments
        """
        with open(filepath, 'w') as f:
            if include_metadata:
                f.write(f"# Experiment: {self.metadata['experiment_name']}\n")
                f.write(f"# Start Time: {self.metadata['start_time']}\n")
                f.write(f"# End Time: {self.metadata['end_time']}\n")
                f.write(f"# Parameters: {self.metadata['parameters']}\n")
                f.write("#\n")

            self.data.to_csv(f, index=False)

    def export_excel(self, filepath: Path):
        """
        Export data to Excel format with formatted sheets.

        Creates a professional Excel workbook with:
        - Data sheet with formatted headers
        - Metadata sheet with experiment information
        - Statistics sheet with calculated statistics
        - Auto-sized columns
        - Header formatting

        Args:
            filepath: Output file path
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # 1. Data sheet
            self.data.to_excel(writer, sheet_name='Data', index=False)
            data_sheet = writer.sheets['Data']

            # Format data headers
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, name='Arial', size=11)

            for col_idx, col in enumerate(self.data.columns, 1):
                cell = data_sheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Auto-size columns
            for col_idx, col in enumerate(self.data.columns, 1):
                column_letter = get_column_letter(col_idx)
                max_length = max(
                    len(str(col)),
                    self.data[col].astype(str).str.len().max() if not self.data.empty else 0
                )
                data_sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

            # 2. Metadata sheet
            meta_df = pd.DataFrame([
                {'Key': k, 'Value': str(v)}
                for k, v in self.metadata.items()
            ])
            meta_df.to_excel(writer, sheet_name='Metadata', index=False)
            meta_sheet = writer.sheets['Metadata']

            # Format metadata headers
            for col_idx in range(1, 3):  # Key and Value columns
                cell = meta_sheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Auto-size metadata columns
            meta_sheet.column_dimensions['A'].width = 20
            meta_sheet.column_dimensions['B'].width = 50

            # 3. Statistics sheet
            stats = self.get_statistics()
            if stats:
                stats_df = pd.DataFrame(stats).T
                stats_df.reset_index(inplace=True)
                stats_df.columns = ['Column'] + list(stats_df.columns[1:])
                stats_df.to_excel(writer, sheet_name='Statistics', index=False)
                stats_sheet = writer.sheets['Statistics']

                # Format statistics headers
                for col_idx in range(1, len(stats_df.columns) + 1):
                    cell = stats_sheet.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Auto-size statistics columns
                for col_idx in range(1, len(stats_df.columns) + 1):
                    stats_sheet.column_dimensions[get_column_letter(col_idx)].width = 15

    def export_json(self, filepath: Path):
        """
        Export data to JSON format.

        Args:
            filepath: Output file path
        """
        output = {
            'metadata': self.metadata,
            'data': self.data.to_dict('records')
        }

        import json
        with open(filepath, 'w') as f:
            json.dump(output, f, indent=2, default=str)

    # Data processing

    def get_statistics(self) -> Dict[str, Any]:
        """
        Calculate basic statistics for the data.

        Returns:
            dict: Statistics for each numeric column
        """
        if self.data.empty:
            return {}

        stats = {}
        for col in self.data.select_dtypes(include=['number']).columns:
            stats[col] = {
                'mean': float(self.data[col].mean()),
                'std': float(self.data[col].std()),
                'min': float(self.data[col].min()),
                'max': float(self.data[col].max()),
                'count': int(self.data[col].count())
            }
        return stats
